from src.core.create_excel import CreateExcel

data = {
        "Command Center": {
            "monthly_income": {
                "primary_income": 6300,
                "secondary_income": 1800,
                "other_income": 2050,
            }
        },
    "Irregular Expense System": {
        "categories": {
            "Home & Property": {
                "Home insurance (annual)":      { "annualCost": 0.0 },
                "Home maintenance & repairs":   { "annualCost": 0.0 }
            },

            "Transport": {
                "Car insurance (annual)":       { "annualCost": 0.0 },
                "Car service / MOT":            { "annualCost": 0.0 },
                "Road tax":                     { "annualCost": 0.0 }
            },

            "Family": {
                "School fees / activities":         { "annualCost": 0.0 },
                "Birthday & Christmas gifts":       { "annualCost": 0.0 }
            },

            "Lifestyle": {
                "Holidays & travel":                { "annualCost": 0.0 },
                "Clothing & seasonal wardrobe":     { "annualCost": 0.0 }
            },

            "Health": {
                "Dental / medical (annual)":        { "annualCost": 0.0 }
            },

            "Other": {
                "Other annual cost 1":              { "annualCost": 0.0 },
                "Other annual cost 2":              { "annualCost": 0.0 },
                "Other annual cost 3":              { "annualCost": 0.0 }
            }
        }
    },

    "Net Position Snapshot": {
        "liquidityReserve": {
            "items": {
                "Cash & Immediate Access Buffers": { "currentValue": 0.0 }
            }
        },

        "wealthVelocityAssets": {
            "items": {
                "Investments (funds, stocks, ETFs)":    { "currentValue": 0.0 },
                "Pension / retirement accounts":        { "currentValue": 0.0 },
                "Property — full market value":         { "currentValue": 0.0 },
                "Other wealth-building assets":         { "currentValue": 0.0 }
            }
        },

        "structuralLiabilities": {
            "items": {
                "Mortgage":                         { "currentValue": 0.0 },
                "Car loan":                         { "currentValue": 0.0 },
                "Student loans":                    { "currentValue": 0.0 },
                "Credit cards & short-term debt":   { "currentValue": 0.0 },
                "Other liabilities":                { "currentValue": 0.0 }
            }
        }
    },

    "Monthly Activation": {}
    }
excel = CreateExcel()
path = excel.update_excel(data=data)
from openpyxl import load_workbook
print(path)



wb = load_workbook(path)

print(wb.sheetnames)


sheet = wb['Command Center']
for row in sheet.iter_rows(values_only=True):
    print(row)