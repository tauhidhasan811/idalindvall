import os
import copy
from openpyxl import load_workbook
from src.cellmaper import get_cell_map
from src.hyperparameter import cell_design


def get_top_left_cell(ws, cell_addr):
    for merged_range in ws.merged_cells.ranges:
        if cell_addr in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return ws[cell_addr]


class CreateExcel:
    def __init__(self, f_name: str = "update"):
        os.makedirs('data/files', exist_ok=True)
        self.file_path = "data/Budget_Method_templete.xlsx"
        self.output_path = f"data/files/{f_name}.xlsx"
        self.budget_method_names = ["Command Center", "Irregular Expense System", "Net Position Snapshot", "Monthly Activation"]


    def update_excel(self, data: dict):
        wb = load_workbook(self.file_path)

        for budget_method_name in self.budget_method_names:
            ws = wb[budget_method_name]

            cell_map = get_cell_map(
                data=data[budget_method_name],
                budget_method_name=budget_method_name
            )

            income_cells = cell_design[budget_method_name]

            for cell_addr, value in cell_map.items():
                cell = get_top_left_cell(ws, cell_addr)

                original_fill = copy.copy(cell.fill)
                original_font = copy.copy(cell.font)
                original_border = copy.copy(cell.border)
                original_alignment = copy.copy(cell.alignment)
                original_protection = copy.copy(cell.protection)
                original_number_format = cell.number_format

                cell.value = value

                cell.fill = original_fill
                cell.font = original_font
                cell.border = original_border
                cell.alignment = original_alignment
                cell.protection = original_protection
                cell.number_format = original_number_format

                if cell_addr in income_cells:
                    cell.number_format = '#,##0 "kr"'

        wb.save(self.output_path)
        print(wb.save(self.output_path))
        print("All cells updated successfully!")
        return self.output_path